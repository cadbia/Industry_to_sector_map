import pandas as pd

# 1) Configuration
FILE_PATH = 'Industry_sector_raw.xlsx'         # your Excel file
INPUT_SHEET = 'Sheet1'                     # sheet with industry & sector columns
OUTPUT_SHEET = 'Industry_sector_map'      # name for the new summary sheet

# 2) Read the data
df = pd.read_excel(FILE_PATH, sheet_name=INPUT_SHEET)

# 3) Count how often each (Industry, Sector) pairing appears
pair_counts = (
    df
    .groupby(['Industry', 'Sector'])
    .size()
    .reset_index(name='count')
)

# 4) Compute the total occurrences per Industry
total_counts = (
    df
    .groupby('Industry')
    .size()
    .reset_index(name='total')
)

# 5) Merge counts so each row has count and total for its Industry
merged = pair_counts.merge(total_counts, on='Industry')

# 6) For each Industry, pick the Sector with the highest count
idx = merged.groupby('Industry')['count'].idxmax()
best = merged.loc[idx].copy()

# 7) Build the “match details” string
def make_detail(row):
    return f"{row['count']}/{row['total']} {row['Industry']} as industry were paired with {row['Sector']} as sector"

best['Match Details'] = best.apply(make_detail, axis=1)

# 8) Prepare the summary DataFrame
summary = best[['Industry', 'Sector', 'Match Details']] \
            .rename(columns={'Sector': 'Most Frequent Sector'})

# 9) Write the summary to a new sheet in the same workbook
from openpyxl import load_workbook

# Check if the output sheet already exists and remove it
try:
    book = load_workbook(FILE_PATH)
    if OUTPUT_SHEET in book.sheetnames:
        std = book[OUTPUT_SHEET]
        book.remove(std)
    book.save(FILE_PATH)
except Exception as e:
    print(f"Warning: Could not modify existing workbook: {e}")

# Write the summary to the new sheet
with pd.ExcelWriter(FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    summary.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)

print(f"✔️  Added sheet '{OUTPUT_SHEET}' with {len(summary)} rows to {FILE_PATH}")
